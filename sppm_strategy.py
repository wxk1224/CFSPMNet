
import csv
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import matplotlib.pyplot as plt
import numpy as np
import torch
import torch.nn.functional as F
from torch.utils.data import DataLoader, Dataset

try:
    from .channel_name_reader import get_channel_names
except ImportError:
    from channel_name_reader import get_channel_names


# ---- hemiplegia_canonicalization.py ----
ROOT = Path(__file__).resolve().parents[2]


def _normalize_side(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"l", "left", "left hand", "left-hand"}:
        return "left"
    if text in {"r", "right", "right hand", "right-hand"}:
        return "right"
    return None


def _normalize_float(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _header_key(value):
    return str(value).strip().lower().replace(" ", "")


def _dataset_key(dataset_name):
    if dataset_name.startswith("XW"):
        return "XW"
    if dataset_name.startswith("TY"):
        return "TY"
    raise ValueError(f"Unsupported dataset for canonicalization: {dataset_name}")


def _flip_pairs_for_dataset(dataset_name):
    key = _dataset_key(dataset_name)
    return XW_FLIP_PAIRS if key == "XW" else TY_FLIP_PAIRS


def build_flip_indices(dataset_name):
    channels = get_channel_names(dataset_name)
    swap_map = {left: right for left, right in _flip_pairs_for_dataset(dataset_name)}
    swap_map.update({right: left for left, right in _flip_pairs_for_dataset(dataset_name)})
    return np.asarray([channels.index(swap_map.get(ch, ch)) for ch in channels], dtype=np.int64)


def _load_xw_metadata(xlsx_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read XW metadata.") from exc

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"XW metadata file is empty: {xlsx_path}")

    headers = {_header_key(name): idx for idx, name in enumerate(rows[0]) if name is not None}
    idx_col = headers.get("idx")
    side_col = headers.get("side")
    dominant_col = headers.get("dominant")
    mrs_col = headers.get("mrs")

    if idx_col is None or side_col is None:
        raise ValueError(f"XW metadata file is missing required columns: {xlsx_path}")

    metadata = {}
    for row in rows[1:]:
        if row[idx_col] is None:
            continue
        subject_id = int(row[idx_col])
        paralysis_side = _normalize_side(row[side_col])
        dominant_hand = _normalize_side(row[dominant_col]) if dominant_col is not None else None
        mrs = _normalize_float(row[mrs_col]) if mrs_col is not None else None
        metadata[subject_id] = {
            "subject_id": subject_id,
            "paralysis_side": paralysis_side,
            "dominant_hand": dominant_hand,
            "mrs": mrs,
            "need_flip": paralysis_side == "left",
        }
    return metadata


def _load_ty_metadata(csv_path):
    metadata = {}
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            subject_id_text = str(row.get("subject_id", "")).strip()
            if not subject_id_text:
                continue
            subject_id = int(subject_id_text)
            paralysis_side = _normalize_side(row.get("paralysis_side"))
            dominant_hand = _normalize_side(row.get("dominant_hand"))
            mrs = _normalize_float(row.get("mrs"))
            metadata[subject_id] = {
                "subject_id": subject_id,
                "paralysis_side": paralysis_side,
                "dominant_hand": dominant_hand,
                "mrs": mrs,
                "need_flip": paralysis_side == "left",
            }
    return metadata


def load_subject_metadata(dataset_name, xw_metadata_path=None, ty_metadata_path=None):
    if dataset_name.startswith("XW"):
        xw_metadata_path = Path(xw_metadata_path or ROOT / "assets" / "XwSubInfo.xlsx")
        if not xw_metadata_path.exists():
            raise FileNotFoundError(f"XW metadata file not found: {xw_metadata_path}")
        return _load_xw_metadata(xw_metadata_path)

    if dataset_name.startswith("TY"):
        ty_metadata_path = Path(ty_metadata_path or ROOT / "assets" / "TySubInfo.csv")
        if not ty_metadata_path.exists():
            raise FileNotFoundError(f"TY metadata file not found: {ty_metadata_path}")
        return _load_ty_metadata(ty_metadata_path)

    raise ValueError(f"Unsupported dataset for metadata loading: {dataset_name}")


def validate_subject_metadata(subject_metadata: Dict[int, Dict], subject_ids: List[int], dataset_name):
    missing_subjects = [sid for sid in subject_ids if sid not in subject_metadata]
    if missing_subjects:
        raise ValueError(f"{dataset_name} metadata is missing subjects: {missing_subjects}")

    invalid_sides = [
        sid
        for sid in subject_ids
        if subject_metadata[sid].get("paralysis_side") not in {"left", "right"}
    ]
    if invalid_sides:
        raise ValueError(
            f"{dataset_name} metadata has missing or invalid paralysis_side for subjects: {invalid_sides}"
        )


def _remap_labels_to_affected_unaffected(y, paralysis_side):
    y = np.asarray(y, dtype=np.int64)
    unique_labels = set(np.unique(y).tolist())
    if not unique_labels.issubset({0, 1}):
        raise ValueError(f"Expected binary labels 0/1, got {sorted(unique_labels)}")

    if paralysis_side == "left":
        affected_mask = y == 0
        unaffected_mask = y == 1
    elif paralysis_side == "right":
        affected_mask = y == 1
        unaffected_mask = y == 0
    else:
        raise ValueError(f"Invalid paralysis_side: {paralysis_side}")

    remapped = np.full_like(y, fill_value=-1)
    remapped[affected_mask] = 0
    remapped[unaffected_mask] = 1
    if np.any(remapped < 0):
        raise ValueError("Failed to remap labels to affected/unaffected.")
    return remapped


def canonicalize_subject_trials(x, y, dataset_name, subject_id, subject_metadata):
    if subject_id not in subject_metadata:
        raise ValueError(f"Subject {subject_id} is missing from metadata.")

    meta = dict(subject_metadata[subject_id])
    paralysis_side = meta.get("paralysis_side")
    if paralysis_side not in {"left", "right"}:
        raise ValueError(f"Subject {subject_id} has invalid paralysis_side: {paralysis_side}")

    x = np.asarray(x, dtype=np.float32)
    y = np.asarray(y, dtype=np.int64)
    if x.ndim != 3:
        raise ValueError(f"Expected subject trials with shape [N, C, T], got {x.shape}")

    canonical_x = x.copy()
    if meta.get("need_flip", False):
        flip_indices = build_flip_indices(dataset_name)
        canonical_x = canonical_x[:, flip_indices, :]

    canonical_y = _remap_labels_to_affected_unaffected(y, paralysis_side)
    meta["label_space"] = {"0": "affected", "1": "unaffected"}
    return canonical_x, canonical_y, meta


# ---- target_adaptation_matching.py ----
def _summarize_matching(probabilities, accepted_mask):
    confidences, _ = torch.max(probabilities, dim=1)
    entropy = -torch.sum(probabilities * torch.log(probabilities + 1e-10), dim=1)
    rejected_mask = ~accepted_mask

    def masked_mean(values, mask):
        if int(mask.sum().item()) == 0:
            return 0.0
        return float(values[mask].mean().item())

    return {
        "accepted_count": int(accepted_mask.sum().item()),
        "rejected_count": int(rejected_mask.sum().item()),
        "accepted_confidence": masked_mean(confidences, accepted_mask),
        "rejected_confidence": masked_mean(confidences, rejected_mask),
        "accepted_entropy": masked_mean(entropy, accepted_mask),
        "rejected_entropy": masked_mean(entropy, rejected_mask),
    }


def apply_confidence_pseudo_label_matching(probabilities, prob_threshold=0.90, existing_pseudo=None):
    if probabilities.ndim != 2:
        raise ValueError(f"Expected probabilities with shape [B, C], got {probabilities.shape}")

    confidences, predictions = torch.max(probabilities, dim=1)
    accepted_mask = confidences >= prob_threshold

    one_hot = torch.zeros_like(probabilities)
    one_hot[accepted_mask, predictions[accepted_mask]] = 1.0

    if existing_pseudo is None:
        updated_pseudo = one_hot
    else:
        updated_pseudo = existing_pseudo.clone()
        updated_pseudo[accepted_mask] = one_hot[accepted_mask]

    summary = _summarize_matching(probabilities, accepted_mask)
    summary.update(
        {
            "updated_pseudo": updated_pseudo,
            "accepted_mask": accepted_mask,
            "predictions": predictions,
            "confidences": confidences,
            "similarities": None,
        }
    )
    return summary


def apply_sppm_signature_prototype_matching(
    probabilities,
    signature_vectors,
    shared_prototypes,
    class_wise_matching_tolerance,
    prob_threshold=0.90,
    existing_pseudo=None,
):
    if probabilities.ndim != 2:
        raise ValueError(f"Expected probabilities with shape [B, C], got {probabilities.shape}")
    if signature_vectors.ndim != 2:
        raise ValueError(f"Expected signature_vectors with shape [B, D], got {signature_vectors.shape}")

    signature_vectors = F.normalize(signature_vectors, p=2, dim=1)
    shared_prototypes = F.normalize(shared_prototypes, p=2, dim=1)

    confidences, predictions = torch.max(probabilities, dim=1)
    predicted_shared_prototypes = shared_prototypes[predictions]
    predicted_matching_tolerance = class_wise_matching_tolerance[predictions]
    similarities = torch.sum(signature_vectors * predicted_shared_prototypes, dim=1)

    accepted_mask = (confidences >= prob_threshold) & (similarities >= predicted_matching_tolerance)

    one_hot = torch.zeros_like(probabilities)
    one_hot[accepted_mask, predictions[accepted_mask]] = 1.0

    if existing_pseudo is None:
        updated_pseudo = one_hot
    else:
        updated_pseudo = existing_pseudo.clone()
        updated_pseudo[accepted_mask] = one_hot[accepted_mask]

    summary = _summarize_matching(probabilities, accepted_mask)
    summary.update(
        {
            "updated_pseudo": updated_pseudo,
            "accepted_mask": accepted_mask,
            "predictions": predictions,
            "confidences": confidences,
            "similarities": similarities,
        }
    )
    return summary


# ---- private_signature_prototypes.py ----
PRIVATE_SIGNATURE_CHANNEL_GROUPS = {
    "XW": {
        "left": ["FC3", "C3", "CP3"],
        "right": ["FC4", "C4", "CP4"],
        "midline": ["FCz", "Cz", "CPz"],
    },
    "TY": {
        "left": ["FC3", "C3", "CP3"],
        "right": ["FC4", "C4", "CP4"],
        "midline": ["FCz", "Cz", "CPz"],
    },
}


def _dataset_key(dataset_name):
    if dataset_name.startswith("XW"):
        return "XW"
    if dataset_name.startswith("TY"):
        return "TY"
    raise ValueError(f"Unsupported dataset for private signature computation: {dataset_name}")


def _resolve_indices(channel_names, signature_channel_names):
    channel_to_index = {name: idx for idx, name in enumerate(channel_names)}
    missing = [name for name in signature_channel_names if name not in channel_to_index]
    if missing:
        raise ValueError(f"Missing private signature channels: {missing}")
    return [channel_to_index[name] for name in signature_channel_names]


def get_private_signature_channel_indices(dataset_name):
    dataset_key = _dataset_key(dataset_name)
    channel_names = get_channel_names(dataset_name)
    signature_channel_names = PRIVATE_SIGNATURE_CHANNEL_GROUPS[dataset_key]
    return {
        "left": _resolve_indices(channel_names, signature_channel_names["left"]),
        "right": _resolve_indices(channel_names, signature_channel_names["right"]),
        "midline": _resolve_indices(channel_names, signature_channel_names["midline"]),
    }


def l2_normalize(vectors, axis=1, eps=1e-8):
    norms = np.linalg.norm(vectors, axis=axis, keepdims=True)
    return vectors / np.clip(norms, eps, None)


def compute_private_signature_features(x, dataset_name):
    x = np.asarray(x, dtype=np.float32)
    if x.ndim != 3:
        raise ValueError(f"Expected input shape [N, C, T], got {x.shape}")

    signature_indices = get_private_signature_channel_indices(dataset_name)

    left_motor = x[:, signature_indices["left"], :].mean(axis=1)
    right_motor = x[:, signature_indices["right"], :].mean(axis=1)
    midline = x[:, signature_indices["midline"], :].mean(axis=1)
    asymmetry = np.abs(left_motor - right_motor)

    signature_vector = np.concatenate([left_motor, right_motor, asymmetry, midline], axis=1)
    signature_vector = l2_normalize(signature_vector, axis=1)

    return {
        "left_motor": left_motor,
        "right_motor": right_motor,
        "midline": midline,
        "asymmetry": asymmetry,
        "signature_vector": signature_vector.astype(np.float32),
    }


def build_shared_private_signature_prototypes(signature_vectors, labels, num_classes=2, floor=0.70):
    signature_vectors = np.asarray(signature_vectors, dtype=np.float32)
    labels = np.asarray(labels, dtype=np.int64)

    shared_prototypes = []
    class_wise_matching_tolerance = []

    for class_id in range(num_classes):
        class_vectors = signature_vectors[labels == class_id]
        if len(class_vectors) == 0:
            raise ValueError(f"No source samples found for class {class_id}.")

        shared_prototype = class_vectors.mean(axis=0, keepdims=True)
        shared_prototype = l2_normalize(shared_prototype, axis=1)[0]

        similarities = class_vectors @ shared_prototype
        matching_tolerance = max(float(similarities.mean() - similarities.std()), float(floor))

        shared_prototypes.append(shared_prototype.astype(np.float32))
        class_wise_matching_tolerance.append(matching_tolerance)

    return np.stack(shared_prototypes, axis=0), np.asarray(class_wise_matching_tolerance, dtype=np.float32)


# ---- reusable SPPM training strategy from train_cfspmnet.py ----
DEFAULT_XW_SUB_LIST = [2, 5, 8, 9, 11, 12, 14, 17, 21, 23, 24, 26, 27, 28, 30, 32, 33, 37, 38, 43, 44, 47, 49, 50]
SUMMARY_METRICS = ["Acc", "Kappa", "F1-Score", "Precision", "Recall", "AUC", "Latency(ms)"]

DATASET_PRESETS = {
    "XW_30Chs": {
        "batch_size": 40,
        "eval_batch_size": 40,
        "lr": 0.001,
        "weight_decay": 0.001,
        "patience": 30,
        "seed": 2,
        "alpha": 0.98,
        "pseudo_threshold": 0.60,
        "matching_tolerance_floor": 0.50,
        "pseudo_warmup_epochs": 25,
    },
    "TY_250hz_new_full": {
        "batch_size": 64,
        "eval_batch_size": 64,
        "lr": 0.001,
        "weight_decay": 0.001,
        "patience": 30,
        "seed": 2,
        "alpha": 0.95,
        "pseudo_threshold": 0.40,
        "matching_tolerance_floor": 0.45,
        "pseudo_warmup_epochs": 10,
    },
}


@dataclass(frozen=True)
class CFSPMNetVariantConfig:
    name: str
    use_target_pseudo: bool
    use_private_signature_matching: bool
    use_entropy_weight: bool
    static_pseudo: bool
    use_frsmamba: bool
    use_temporal_encoder: bool


VARIANT_CONFIGS = {
    "CFSPMNetSourceOnly": CFSPMNetVariantConfig("CFSPMNetSourceOnly", False, False, False, True, True, True),
    "CFSPMNetConfidenceOnly": CFSPMNetVariantConfig("CFSPMNetConfidenceOnly", True, False, True, False, True, True),
    "CFSPMNetSPPM": CFSPMNetVariantConfig("CFSPMNetSPPM", True, True, True, False, True, True),
    "CFSPMNetNoSPPM": CFSPMNetVariantConfig("CFSPMNetNoSPPM", True, False, True, False, True, True),
    "CFSPMNet": CFSPMNetVariantConfig("CFSPMNet", True, True, False, False, True, True),
    "CFSPMNetStaticTarget": CFSPMNetVariantConfig("CFSPMNetStaticTarget", True, True, True, True, True, True),
    "CFSPMNetNoFRSMamba": CFSPMNetVariantConfig("CFSPMNetNoFRSMamba", True, True, True, False, False, True),
    "CFSPMNetTokenizerOnly": CFSPMNetVariantConfig("CFSPMNetTokenizerOnly", True, True, True, False, False, False),
}


class SourceDataset(Dataset):
    def __init__(self, x, y):
        self.x = torch.tensor(x, dtype=torch.float32)
        self.y = torch.tensor(y, dtype=torch.long)

    def __len__(self):
        return len(self.x)

    def __getitem__(self, idx):
        return self.x[idx], self.y[idx]


class TargetPseudoDataset(Dataset):
    def __init__(self, x, signature_vectors, num_classes):
        self.x = torch.tensor(x, dtype=torch.float32)
        self.signature_vectors = torch.tensor(signature_vectors, dtype=torch.float32)
        self.pseudo_labels = torch.zeros((len(x), num_classes), dtype=torch.float32)

    def update_pseudo_labels(self, indices, new_labels):
        self.pseudo_labels[indices] = new_labels

    def active_ratio(self):
        return float((self.pseudo_labels.sum(dim=1) > 0).float().mean().item())

    def __len__(self):
        return len(self.x)

    def __getitem__(self, idx):
        return self.x[idx], self.pseudo_labels[idx], self.signature_vectors[idx], idx


class TargetEvalDataset(Dataset):
    def __init__(self, x, y):
        self.x = torch.tensor(x, dtype=torch.float32)
        self.y = torch.tensor(y, dtype=torch.long)

    def __len__(self):
        return len(self.x)

    def __getitem__(self, idx):
        return self.x[idx], self.y[idx]

def build_dynamic_dataloaders(source_dataset, target_train_dataset, target_eval_dataset, batch_size, num_workers):
    source_batch = min(batch_size, len(source_dataset))
    target_batch = min(batch_size, len(target_train_dataset))
    source_drop_last = len(source_dataset) > 1 and (len(source_dataset) % source_batch == 1)
    target_drop_last = len(target_train_dataset) > 1 and (len(target_train_dataset) % target_batch == 1)

    source_loader = DataLoader(
        source_dataset,
        batch_size=source_batch,
        shuffle=True,
        drop_last=source_drop_last,
        num_workers=num_workers,
        pin_memory=torch.cuda.is_available(),
    )
    target_train_loader = DataLoader(
        target_train_dataset,
        batch_size=target_batch,
        shuffle=True,
        drop_last=target_drop_last,
        num_workers=num_workers,
        pin_memory=torch.cuda.is_available(),
    )
    target_eval_loader = DataLoader(
        target_eval_dataset,
        batch_size=target_batch,
        shuffle=False,
        drop_last=False,
        num_workers=num_workers,
        pin_memory=torch.cuda.is_available(),
    )
    target_init_loader = DataLoader(
        target_train_dataset,
        batch_size=target_batch,
        shuffle=False,
        drop_last=False,
        num_workers=num_workers,
        pin_memory=torch.cuda.is_available(),
    )
    return source_loader, target_train_loader, target_eval_loader, target_init_loader

def parameter_count(model):
    return sum(param.numel() for param in model.parameters() if param.requires_grad)


def sppm_matching_function(args, probabilities, signature_vectors, shared_prototypes, class_wise_matching_tolerance, existing_pseudo):
    if args.use_private_signature_matching:
        return apply_sppm_signature_prototype_matching(
            probabilities=probabilities,
            signature_vectors=signature_vectors,
            shared_prototypes=shared_prototypes,
            class_wise_matching_tolerance=class_wise_matching_tolerance,
            prob_threshold=args.pseudo_threshold,
            existing_pseudo=existing_pseudo,
        )
    return apply_confidence_pseudo_label_matching(
        probabilities=probabilities,
        prob_threshold=args.pseudo_threshold,
        existing_pseudo=existing_pseudo,
    )


def init_matching_aggregator():
    return {
        "accepted_count": 0,
        "rejected_count": 0,
        "accepted_conf_sum": 0.0,
        "rejected_conf_sum": 0.0,
        "accepted_entropy_sum": 0.0,
        "rejected_entropy_sum": 0.0,
    }


def accumulate_matching_stats(agg, matching_result):
    accepted_count = int(matching_result["accepted_count"])
    rejected_count = int(matching_result["rejected_count"])
    agg["accepted_count"] += accepted_count
    agg["rejected_count"] += rejected_count
    agg["accepted_conf_sum"] += accepted_count * float(matching_result["accepted_confidence"])
    agg["rejected_conf_sum"] += rejected_count * float(matching_result["rejected_confidence"])
    agg["accepted_entropy_sum"] += accepted_count * float(matching_result["accepted_entropy"])
    agg["rejected_entropy_sum"] += rejected_count * float(matching_result["rejected_entropy"])


def finalize_matching_stats(agg):
    accepted_count = agg["accepted_count"]
    rejected_count = agg["rejected_count"]
    return {
        "accepted_count": accepted_count,
        "rejected_count": rejected_count,
        "accepted_confidence": agg["accepted_conf_sum"] / max(accepted_count, 1),
        "rejected_confidence": agg["rejected_conf_sum"] / max(rejected_count, 1),
        "accepted_entropy": agg["accepted_entropy_sum"] / max(accepted_count, 1),
        "rejected_entropy": agg["rejected_entropy_sum"] / max(rejected_count, 1),
    }


def save_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def plot_ratio_curves(csv_rows, output_path, title):
    epochs = [int(row["epoch"]) for row in csv_rows]
    accepted = [float(row["accepted_ratio"]) for row in csv_rows]
    active = [float(row["active_ratio"]) for row in csv_rows]
    plt.figure(figsize=(7, 4))
    plt.plot(epochs, accepted, label="accepted_ratio", linewidth=1.6)
    plt.plot(epochs, active, label="active_ratio", linewidth=1.6)
    plt.xlabel("Epoch")
    plt.ylabel("Ratio")
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=160)
    plt.close()


def save_final_report(args, total_time, all_metrics):
    report_path = args.result_dir / f"{args.exp_name}_{args.model}_result.txt"
    total_seconds = int(total_time)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    with report_path.open("w", encoding="utf-8") as handle:
        handle.write("Experiment Report\n")
        handle.write(f"Model: {args.model}\n")
        handle.write(f"Variant: {args.variant}\n")
        handle.write(f"Dataset: {args.dataset}\n")
        handle.write("Protocol: LOSO Ablation\n")
        handle.write(f"Pseudo Threshold: {args.pseudo_threshold:.2f}\n")
        handle.write(f"Pseudo Warmup Epochs: {args.pseudo_warmup_epochs}\n")
        handle.write(f"Class-wise Matching Tolerance Floor: {args.matching_tolerance_floor:.2f}\n")
        handle.write(f"Rhythm Branch Mode: {args.rhythm_branch_mode}\n")
        handle.write(f"Rhythm Impl: {args.rhythm_impl}\n")
        handle.write(f"Fourier Rhythm Sparsity Threshold: {args.rhythm_sparsity_threshold:.4f}\n")
        handle.write(f"Rhythm Low Ratio: {args.rhythm_low_ratio:.4f}\n")
        handle.write(f"Freq Num Blocks: {args.rhythm_num_blocks if args.rhythm_num_blocks is not None else 'auto'}\n")
        handle.write(f"Init Checkpoint Root: {args.init_ckpt_root or 'none'}\n")
        handle.write(f"Date: {args.exp_time}\n")
        handle.write(f"Total Training Time: {hours:02d}:{minutes:02d}:{seconds:02d}\n")
        handle.write("=" * 80 + "\n")
        for metric_name in SUMMARY_METRICS:
            values = np.asarray(all_metrics[metric_name], dtype=np.float64)
            handle.write(f"{metric_name:<12}: {values.mean():.4f} +/- {values.std():.4f}\n")
        handle.write("=" * 80 + "\n")
    return report_path


def find_init_checkpoint(init_ckpt_root: Path | None, subject_id: int, fold_idx: int) -> Path | None:
    if init_ckpt_root is None or not init_ckpt_root.exists():
        return None

    patterns = [
        f"sub{subject_id}_fold{fold_idx}_*.pth",
        f"sub{subject_id}_*.pth",
    ]
    for pattern in patterns:
        matches = sorted(init_ckpt_root.glob(pattern), key=lambda path: path.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    return None


def load_init_weights_if_available(model, checkpoint_path: Path | None, device, logger):
    if checkpoint_path is None:
        return

    state_dict = torch.load(checkpoint_path, map_location=device, weights_only=True)
    load_result = model.load_state_dict(state_dict, strict=False)
    missing = list(load_result.missing_keys)
    unexpected = list(load_result.unexpected_keys)
    logger.print(f"Warm start from {checkpoint_path.name}")
    if missing:
        logger.print(f"Missing keys: {len(missing)}")
    if unexpected:
        logger.print(f"Unexpected keys: {len(unexpected)}")


def apply_freeze_strategy(args, model, logger):
    frozen_groups = []

    if args.freeze_physiological_tokenizer:
        for param in model.physiological_tokenizer.parameters():
            param.requires_grad = False
        frozen_groups.append("physiological_tokenizer")

    if args.freeze_position_encoding:
        for param in model.position_encoding.parameters():
            param.requires_grad = False
        frozen_groups.append("position_encoding")

    if args.freeze_prediction_head:
        for param in model.prediction_head.parameters():
            param.requires_grad = False
        frozen_groups.append("prediction_head")

    if args.freeze_nonrhythmic_temporal and hasattr(model, "temporal_encoder"):
        for block in model.temporal_encoder:
            if not hasattr(block, "temporal_mixer"):
                continue
            for name, param in block.temporal_mixer.named_parameters():
                if not name.startswith("fourier_rhythmic_modulator."):
                    param.requires_grad = False
            if hasattr(block, "norm"):
                for param in block.norm.parameters():
                    param.requires_grad = False
        frozen_groups.append("nonrhythmic_temporal")

    if args.freeze_fourier_rhythmic_state_modeling and hasattr(model, "temporal_encoder"):
        for block in model.temporal_encoder:
            if hasattr(block, "rhythmic_state_modeling") and block.rhythmic_state_modeling is not None:
                for param in block.rhythmic_state_modeling.parameters():
                    param.requires_grad = False
        frozen_groups.append("rhythmic_state_modeling")

    if frozen_groups:
        logger.print(f"Frozen modules: {', '.join(frozen_groups)}")

def initialize_target_pseudo_labels(args, model, target_init_loader, shared_prototypes, class_wise_matching_tolerance):
    model.eval()
    total_accepted = 0
    total_samples = 0
    matching_agg = init_matching_aggregator()

    with torch.no_grad():
        for x_t, pseudo_y, signature_t, idx_t in target_init_loader:
            x_t = x_t.to(args.device, non_blocking=True)
            pseudo_y = pseudo_y.to(args.device, non_blocking=True)
            signature_t = signature_t.to(args.device, non_blocking=True)

            logits = model(x_t)
            probs = torch.softmax(logits, dim=1)
            matching_result = sppm_matching_function(args, probs, signature_t, shared_prototypes, class_wise_matching_tolerance, pseudo_y)
            target_init_loader.dataset.update_pseudo_labels(idx_t, matching_result["updated_pseudo"].cpu())
            total_accepted += int(matching_result["accepted_count"])
            total_samples += len(idx_t)
            accumulate_matching_stats(matching_agg, matching_result)

    return total_accepted / max(total_samples, 1), finalize_matching_stats(matching_agg)


def train_source_only_epoch(source_loader, model, optimizer, device):
    model.train()
    total_loss_epoch = 0.0
    total_batches = len(source_loader)
    if total_batches == 0:
        return 0.0

    for x_s, y_s in source_loader:
        x_s = x_s.to(device, non_blocking=True)
        y_s = y_s.to(device, non_blocking=True)
        optimizer.zero_grad(set_to_none=True)
        src_logits = model(x_s)
        loss_source = F.cross_entropy(src_logits, y_s)
        loss_source.backward()
        optimizer.step()
        total_loss_epoch += loss_source.item()

    return total_loss_epoch / total_batches


def train_sppm_epoch(args, source_loader, target_loader, model, optimizer, shared_prototypes, class_wise_matching_tolerance):
    model.train()
    total_loss_epoch = 0.0
    total_source_loss = 0.0
    total_target_loss = 0.0
    total_batches = len(source_loader)
    matching_agg = init_matching_aggregator()

    if total_batches == 0:
        return {
            "train_loss": 0.0,
            "source_loss": 0.0,
            "target_loss": 0.0,
            "accepted_ratio": 0.0,
            "active_ratio": 0.0,
            "matching_stats": finalize_matching_stats(matching_agg),
        }

    target_iter = iter(target_loader)
    total_target_samples = 0

    for x_s, y_s in source_loader:
        try:
            x_t, tgt_pseudo_y, signature_t, tgt_idx = next(target_iter)
        except StopIteration:
            target_iter = iter(target_loader)
            x_t, tgt_pseudo_y, signature_t, tgt_idx = next(target_iter)

        x_s = x_s.to(args.device, non_blocking=True)
        y_s = y_s.to(args.device, non_blocking=True)
        x_t = x_t.to(args.device, non_blocking=True)
        tgt_pseudo_y = tgt_pseudo_y.to(args.device, non_blocking=True)
        signature_t = signature_t.to(args.device, non_blocking=True)

        optimizer.zero_grad(set_to_none=True)

        src_logits = model(x_s)
        loss_source = F.cross_entropy(src_logits, y_s)

        tgt_logits = model(x_t)
        log_probs = F.log_softmax(tgt_logits, dim=1)
        target_weights = 1.0
        if args.use_entropy_weight:
            tgt_probs = torch.softmax(tgt_logits, dim=1)
            entropy = -torch.sum(tgt_probs * torch.log(tgt_probs + 1e-10), dim=1)
            target_weights = torch.exp(-entropy).detach()
        loss_target_sample = -torch.sum(tgt_pseudo_y * log_probs, dim=1)
        if torch.is_tensor(target_weights):
            loss_target = torch.mean(target_weights * loss_target_sample)
        else:
            loss_target = torch.mean(loss_target_sample)

        total_loss = args.alpha * loss_source + (1.0 - args.alpha) * loss_target
        total_loss.backward()
        optimizer.step()

        if not args.static_pseudo:
            with torch.no_grad():
                updated_logits = model(x_t)
                updated_probs = torch.softmax(updated_logits, dim=1)
                matching_result = sppm_matching_function(args, updated_probs, signature_t, shared_prototypes, class_wise_matching_tolerance, tgt_pseudo_y)
                target_loader.dataset.update_pseudo_labels(tgt_idx, matching_result["updated_pseudo"].cpu())
                accumulate_matching_stats(matching_agg, matching_result)
                total_target_samples += len(tgt_idx)
        else:
            total_target_samples += len(tgt_idx)

        total_loss_epoch += total_loss.item()
        total_source_loss += loss_source.item()
        total_target_loss += loss_target.item()

    matching_stats = finalize_matching_stats(matching_agg)
    return {
        "train_loss": total_loss_epoch / total_batches,
        "source_loss": total_source_loss / total_batches,
        "target_loss": total_target_loss / total_batches,
        "accepted_ratio": matching_stats["accepted_count"] / max(total_target_samples, 1),
        "active_ratio": target_loader.dataset.active_ratio(),
        "matching_stats": matching_stats,
    }


def evaluate_target(model, target_loader, device, n_classes):
    model.eval()
    total_loss = 0.0
    total_samples = 0
    y_true_list = []
    y_pred_list = []
    y_prob_list = []

    with torch.no_grad():
        for x_t, y_t in target_loader:
            x_t = x_t.to(device, non_blocking=True)
            y_t = y_t.to(device, non_blocking=True)
            logits = model(x_t)
            probs = torch.softmax(logits, dim=1)
            preds = logits.argmax(dim=1)
            loss = F.cross_entropy(logits, y_t, reduction="sum")
            total_loss += loss.item()
            total_samples += y_t.size(0)
            y_true_list.append(y_t.cpu().numpy())
            y_pred_list.append(preds.cpu().numpy())
            y_prob_list.append(probs.cpu().numpy())

    y_true = np.concatenate(y_true_list, axis=0)
    y_pred = np.concatenate(y_pred_list, axis=0)
    y_prob = np.concatenate(y_prob_list, axis=0)
    metrics = compute_metrics(y_true=y_true, y_pred=y_pred, y_prob=y_prob, n_classes=n_classes)
    return {
        "val_loss": total_loss / max(total_samples, 1),
        "val_acc": metrics["Acc"],
        "metrics": metrics,
    }
